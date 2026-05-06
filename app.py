"""
Salarium Batch Simulator
========================
Interface Streamlit pour lancer des simulations Salarium sur toutes les
combinaisons de paramètres complémentaires × tous les âges.

Usage : streamlit run app.py
"""

import asyncio
import io
from datetime import datetime
from itertools import product

import pandas as pd
import streamlit as st

from salarium_scraper import run_simulations, Combination
from salarium_options import (
    BRANCHES, REGIONS, PROFESSIONS,
    POSITIONS, FORMATIONS, SEXES, NATIONALITES,
    TAILLES_ENTREPRISE, OUI_NON, TYPES_CONTRAT,
)

MAX_COMBINATIONS = 20
AGE_MAX = 65

st.set_page_config(page_title="Salarium Batch", page_icon="💰", layout="wide")
st.title("💰 Salarium — Simulations multi-combinaisons")

# ---------------------------------------------------------------------------
# Champs principaux (single-select)
# ---------------------------------------------------------------------------
st.subheader("Paramètres principaux")

col1, col2, col3 = st.columns(3)
with col1:
    branche = st.selectbox(
        "Branche économique",
        options=BRANCHES,
        index=BRANCHES.index("66. Activités auxiliaires de services financiers et d'assurance"),
    )
with col2:
    region = st.selectbox(
        "Région",
        options=REGIONS,
        index=0,
    )
with col3:
    profession = st.selectbox(
        "Groupe de professions",
        options=PROFESSIONS,
        index=0,
    )

# ---------------------------------------------------------------------------
# Paramètres numériques
# ---------------------------------------------------------------------------
st.subheader("Âge & horaire")

col_a, col_b = st.columns(2)
with col_a:
    age_start = st.slider(
        "Âge de début (= 0 année de service)",
        min_value=18, max_value=50, value=25, step=1,
        help=f"L'âge boucle de cette valeur jusqu'à {AGE_MAX}. "
             f"Les années de service incrémentent en parallèle (0, 1, 2, …).",
    )
with col_b:
    horaire_hebdo = st.slider(
        "Horaire hebdomadaire (h)",
        min_value=8, max_value=50, value=40, step=1,
        help="Valeur fixe pendant toute la simulation.",
    )

n_ages = AGE_MAX - age_start + 1
st.caption(f"→ {n_ages} simulations d'âge ({age_start} → {AGE_MAX})")

# ---------------------------------------------------------------------------
# Paramètres complémentaires (multi-select)
# ---------------------------------------------------------------------------
st.subheader("Paramètres complémentaires (multi-sélection → combinaisons)")

c1, c2 = st.columns(2)
with c1:
    sel_positions = st.multiselect("Position professionnelle", POSITIONS, default=[POSITIONS[0]])
    sel_formations = st.multiselect("Formation", FORMATIONS, default=[FORMATIONS[0]])
    sel_sexes = st.multiselect("Sexe", SEXES, default=["Homme"])
    sel_nationalites = st.multiselect("Nationalité / Permis", NATIONALITES, default=[NATIONALITES[0]])

with c2:
    sel_tailles = st.multiselect("Taille de l'entreprise", TAILLES_ENTREPRISE, default=[TAILLES_ENTREPRISE[2]])
    sel_treiziemes = st.multiselect("13e salaire", OUI_NON, default=["Oui"])
    sel_paiements = st.multiselect("Paiements spéciaux", OUI_NON, default=["Oui"])
    sel_contrats = st.multiselect("Type de contrat", TYPES_CONTRAT, default=[TYPES_CONTRAT[0]])

# ---------------------------------------------------------------------------
# Compteur de combinaisons
# ---------------------------------------------------------------------------
multi_lists = [
    sel_positions, sel_formations, sel_sexes, sel_nationalites,
    sel_tailles, sel_treiziemes, sel_paiements, sel_contrats,
]
n_combinations = 1
for lst in multi_lists:
    n_combinations *= max(len(lst), 0)

n_total = n_combinations * n_ages
duration_min = (n_total * 4) / 60  # ~4s par simulation

empty_field = any(len(lst) == 0 for lst in multi_lists)

if empty_field:
    st.error("⚠️ Sélectionne au moins une option dans chaque catégorie complémentaire.")
elif n_combinations > MAX_COMBINATIONS:
    st.error(
        f"⚠️ {n_combinations} combinaisons demandées (max autorisé : {MAX_COMBINATIONS}). "
        "Réduis tes sélections."
    )
else:
    st.info(
        f"**{n_combinations} combinaison(s) × {n_ages} âges = {n_total} simulations** — "
        f"durée estimée : ~{duration_min:.1f} min"
    )

# ---------------------------------------------------------------------------
# Lancement
# ---------------------------------------------------------------------------
with st.sidebar:
    st.header("Options")
    url = st.text_input("URL Salarium", value="https://www.salarium.bfs.admin.ch/")
    headless = st.checkbox("Mode headless", value=False)
    delay = st.slider("Délai entre simulations (s)", 0.5, 3.0, 1.0, 0.5)

if "results_df" not in st.session_state:
    st.session_state.results_df = None

run_disabled = empty_field or n_combinations > MAX_COMBINATIONS or n_combinations == 0
run = st.button("🚀 Lancer", type="primary", use_container_width=True, disabled=run_disabled)

if run:
    # Construire la liste des combinaisons (produit cartésien)
    combos = []
    for pos, form, sex, nat, taille, tre, pai, contrat in product(
        sel_positions, sel_formations, sel_sexes, sel_nationalites,
        sel_tailles, sel_treiziemes, sel_paiements, sel_contrats,
    ):
        combos.append(Combination(
            branche=branche, region=region, profession=profession,
            position=pos, formation=form, sexe=sex, nationalite=nat,
            taille=taille, treizieme=tre, paiements=pai, type_contrat=contrat,
            horaire_hebdo=float(horaire_hebdo), age_start=age_start,
        ))

    st.session_state.results_df = None
    progress_bar = st.progress(0.0, text="Initialisation…")
    status_box = st.empty()

    import time
    start_time = time.time()
    last_idx = [0]  # mutable container pour closure

    def on_progress(idx: int, total: int, message: str):
        pct = idx / total if total else 0
        progress_bar.progress(min(pct, 1.0), text=f"[{idx}/{total}] {message[:80]}")

        # ETA basé sur le rythme moyen depuis le début
        elapsed = time.time() - start_time
        if idx > 0 and idx > last_idx[0]:
            last_idx[0] = idx
            avg_per_sim = elapsed / idx
            remaining = (total - idx) * avg_per_sim
            mins = int(remaining // 60)
            secs = int(remaining % 60)
            elapsed_min = int(elapsed // 60)
            elapsed_sec = int(elapsed % 60)
            status_box.info(
                f"⏱️ Écoulé : **{elapsed_min}m {elapsed_sec:02d}s** | "
                f"ETA : **{mins}m {secs:02d}s** restantes | "
                f"Vitesse : {avg_per_sim:.1f}s/simulation"
            )

    try:
        results = asyncio.run(
            run_simulations(
                url=url.strip(),
                combinations=combos,
                age_min=age_start,
                age_max=AGE_MAX,
                headless=headless,
                delay_seconds=float(delay),
                progress_callback=on_progress,
            )
        )
        st.session_state.results_df = pd.DataFrame(results)
        progress_bar.progress(1.0, text="Terminé !")
        elapsed_total = time.time() - start_time
        mins = int(elapsed_total // 60)
        secs = int(elapsed_total % 60)
        status_box.empty()
        st.success(f"✅ {len(results)} simulations complétées en **{mins}m {secs:02d}s**.")
    except Exception as e:
        progress_bar.empty()
        status_box.empty()
        st.error(f"❌ {e}")


# ---------------------------------------------------------------------------
# Construction Excel
# ---------------------------------------------------------------------------
def build_excel(df: pd.DataFrame) -> bytes:
    cols_order = [
        "branche", "region", "profession",
        "position", "formation", "sexe", "nationalite", "taille",
        "treizieme", "paiements", "type_contrat", "horaire_hebdo",
        "age", "annees_service",
        "q1", "mediane", "q3",
    ]
    headers = [
        "Branche", "Région", "Profession",
        "Position", "Formation", "Sexe", "Nationalité", "Taille entreprise",
        "13e salaire", "Paiements spéciaux", "Type de contrat", "Horaire hebdo (h)",
        "Âge", "Années service",
        "25% gagnent moins de (CHF)", "Médiane (CHF)", "25% gagnent plus de (CHF)",
    ]

    df_out = df[cols_order].copy()
    df_out.columns = headers

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_out.to_excel(writer, sheet_name="Résultats", index=False)
        ws = writer.sheets["Résultats"]

        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        n_cols = len(headers)
        n_rows = len(df_out)

        # En-tête
        header_fill = PatternFill("solid", fgColor="2C6E6A")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(top=thin, bottom=thin, left=thin, right=thin)

        for c in range(1, n_cols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 38

        # Données
        for r in range(2, n_rows + 2):
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                cell.alignment = Alignment(horizontal="left" if c <= 11 else "center",
                                           vertical="center", wrap_text=(c <= 11))
                # Format CHF pour les 3 dernières colonnes
                if c >= n_cols - 2 and cell.value is not None:
                    cell.number_format = '#,##0" CHF"'
            if (r - 2) % 2 == 1:
                for c in range(1, n_cols + 1):
                    if ws.cell(row=r, column=c).fill.fgColor.rgb in (None, "00000000"):
                        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="F5F5F5")

        # Largeurs
        widths = [40, 30, 40, 25, 30, 10, 25, 22, 12, 18, 18, 14, 8, 14, 22, 18, 22]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Filtres + freeze
        last_col = get_column_letter(n_cols)
        ws.auto_filter.ref = f"A1:{last_col}{n_rows + 1}"
        ws.freeze_panes = "A2"

    return output.getvalue()


# ---------------------------------------------------------------------------
# Résultats
# ---------------------------------------------------------------------------
if st.session_state.results_df is not None and not st.session_state.results_df.empty:
    st.divider()
    df = st.session_state.results_df

    # Tableau
    st.subheader("Résultats")
    st.dataframe(df, use_container_width=True, hide_index=True, height=350)

    # Graph
    if df["mediane"].notna().any():
        st.subheader("Évolution par âge")
        # Si plusieurs combinaisons, on affiche la moyenne par âge
        if df.groupby(["sexe", "treizieme", "type_contrat"]).ngroups > 1:
            chart_df = df.groupby("age")[["q1", "mediane", "q3"]].mean()
        else:
            chart_df = df.set_index("age")[["q1", "mediane", "q3"]]
        chart_df = chart_df.rename(columns={
            "q1": "25% gagnent moins de",
            "mediane": "Médiane",
            "q3": "25% gagnent plus de",
        })
        st.line_chart(chart_df)

    # Export
    try:
        excel_bytes = build_excel(df)
        st.download_button(
            "⬇️ Télécharger Excel",
            data=excel_bytes,
            file_name=f"salarium_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )
    except Exception as e:
        st.error(f"Erreur Excel : {e}")
